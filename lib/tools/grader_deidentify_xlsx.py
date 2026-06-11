#!/usr/bin/env python3
"""
grader_deidentify_xlsx.py — FERPA-safe Excel workbook audit for the grader.

Part of the canvas-toolbox generic grader skill (v1.0). See:
  - grading_readme.md (faculty-facing pipeline + canonical layout)
  - lib/agents/canvas_grader.md (agent-facing pipeline)

WHAT IT DOES
  Excel submissions (.xlsx) are common in accounting / finance / business /
  data-analytics courses. Sending the raw binary to a grading LLM is wrong
  for two reasons:

    1. **The LLM can't reason about it** — binary blobs don't help; what
       matters is structure (sheets, formulas, formatting, layout, charts).
    2. **FERPA** — the workbook's file properties (`/Author`, `/Last
       Modified By`) typically carry the student's name. The cell content
       can carry it too (a "Name:" header cell is common in templates).

  Instead this adapter generates a plain-text **workbook audit** describing
  the structure of the spreadsheet in a form the LLM can grade against:

  - Sheet names
  - Per sheet: freeze-panes location, used range, column-level formatting
    (column width, bold/font/fill applied to whole columns)
  - Cell details for the first N rows (default 10): address, value
    (truncated), formatting
  - **Formulas, grouped by column with range** (e.g. `B2:B50: =A2*1.1`) —
    so a 50-row tax calculation shows as one summary line, not 50 rows
  - Charts: type and title
  - Tables: name and range

  The audit gets wrapped with a header marker `=== EXCEL WORKBOOK AUDIT ===`
  so the grader prompt can switch to spreadsheet-grading mode (rubric
  evaluates structure, formulas, formatting, layout — not prose).

FERPA BOUNDARY
  - File properties (/Author, /Last Modified By, /Created, /Modified) are
    SCRUBBED from the audit. They never appear in submissions_deid/.
  - Cell values are scrubbed against .known_names.txt + the standard
    secret/email patterns.
  - Console prints keys + counts ONLY, never a name.

GENERALIZED FROM: UniversalGrader's workbook-audit pattern (BYUI colleague's
tool, surveyed 2026-06-11). Their docx documents the architecture; this
is the canvas-toolbox parallel.

DEPENDENCIES
  Requires `openpyxl` (in pyproject.toml deps as of v0.33+). Lazy import
  with a clear install hint if missing.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from __toolbox_version__ import __version__
except ImportError:
    __version__ = "0.0.0+unknown"

from grader_deidentify_databricks import (  # noqa: E402
    EMAIL_RE,
    USERPATH_RE,
    SECRET_PREFIX_RE,
    SECRET_ASSIGN_RE,
    key_for,
)

_MAX_HEAD_ROWS = 10   # cell-level detail for the first N rows of each sheet
_CELL_VALUE_MAX = 60  # truncate cell values for the audit
_MAX_FORMULA_GROUPS = 200  # safety cap on the formula summary
_AUDIT_HEADER = "=== EXCEL WORKBOOK AUDIT ==="


def _col_letter(col_idx: int) -> str:
    """1-indexed column number → letter (A, B, …, AA)."""
    s = ""
    while col_idx:
        col_idx, r = divmod(col_idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _truncate(v: object) -> str:
    s = "" if v is None else str(v)
    if len(s) > _CELL_VALUE_MAX:
        return s[:_CELL_VALUE_MAX] + "…"
    return s


def _summarize_column_runs(formula_rows: list[tuple[str, int, str]]) -> list[str]:
    """Given (col_letter, row, formula) triples, group consecutive same-formula-shape
    rows into ranges. Example: (B,2,'=A2*1.1'), (B,3,'=A3*1.1'), … → 'B2:B50: =A2*1.1'.

    'Same formula shape' here means equal after stripping per-row index numbers
    (e.g. '=A2*1.1' and '=A3*1.1' share shape '=A*1.1' after stripping)."""
    if not formula_rows:
        return []
    formula_rows.sort(key=lambda x: (x[0], x[1]))
    out: list[str] = []
    current = None  # (col, start_row, end_row, sample_formula, shape)

    def shape_of(f: str) -> str:
        # Strip the row number from cell refs like A2, B$10, $C$3 → A, B$, $C$
        return re.sub(r"(\$?[A-Z]+\$?)\d+", r"\1", f)

    for col, row, formula in formula_rows:
        shp = shape_of(formula)
        if current and current[0] == col and current[4] == shp and current[2] + 1 == row:
            current = (col, current[1], row, current[3], shp)
        else:
            if current:
                c, s, e, samp, _ = current
                out.append(f"{c}{s}:{c}{e}: {samp}" if s != e else f"{c}{s}: {samp}")
                if len(out) >= _MAX_FORMULA_GROUPS:
                    return out + [f"… ({len(formula_rows) - len(out)} more formula groups omitted)"]
            current = (col, row, row, formula, shp)
    if current:
        c, s, e, samp, _ = current
        out.append(f"{c}{s}:{c}{e}: {samp}" if s != e else f"{c}{s}: {samp}")
    return out


def build_audit(path: Path) -> tuple[str, list[str]]:
    """Generate the plain-text workbook audit. Returns (audit_text, raw_name_strings)
    where raw_name_strings are cell values that could be name-like (for scrubbing)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Missing dependency: openpyxl. Install via `uv add openpyxl` "
              "or `uv sync` (it's in pyproject.toml deps).", file=sys.stderr)
        raise

    wb = load_workbook(filename=str(path), data_only=False, read_only=False)
    lines: list[str] = [_AUDIT_HEADER, ""]
    lines.append(f"Sheets: {wb.sheetnames}")
    lines.append("")
    cell_value_candidates: list[str] = []  # for name scrubbing

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"## Sheet: {sheet_name}")
        if ws.freeze_panes:
            lines.append(f"Freeze panes: {ws.freeze_panes}")
        lines.append(f"Used range: {ws.dimensions}  (rows {ws.max_row}, cols {ws.max_column})")

        # Column-level formatting (where a column has explicit width or
        # whole-column formatting applied — openpyxl reports this in column_dimensions)
        col_fmt_lines: list[str] = []
        for col_letter, col_dim in ws.column_dimensions.items():
            details = []
            if col_dim.width is not None:
                details.append(f"width={col_dim.width:.1f}")
            if col_dim.hidden:
                details.append("hidden")
            if details:
                col_fmt_lines.append(f"  {col_letter}: {', '.join(details)}")
        if col_fmt_lines:
            lines.append("Column formatting (non-default):")
            lines.extend(col_fmt_lines[:30])  # cap to first 30 columns

        # Cell details for the first N rows
        if ws.max_row > 0 and ws.max_column > 0:
            lines.append("")
            lines.append(f"Cell details (rows 1–{min(_MAX_HEAD_ROWS, ws.max_row)}):")
            for row_idx in range(1, min(_MAX_HEAD_ROWS, ws.max_row) + 1):
                for col_idx in range(1, min(20, ws.max_column) + 1):  # cap at 20 cols/row
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is None:
                        continue
                    val_str = _truncate(cell.value)
                    cell_value_candidates.append(str(cell.value))
                    fmt = []
                    if cell.font and cell.font.bold:
                        fmt.append("bold")
                    if cell.font and cell.font.name and cell.font.name != "Calibri":
                        fmt.append(f"font={cell.font.name}")
                    if (cell.fill and cell.fill.fgColor and
                            cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000"):
                        fmt.append(f"fill={cell.fill.fgColor.rgb}")
                    if cell.alignment and cell.alignment.horizontal:
                        fmt.append(f"align={cell.alignment.horizontal}")
                    fmt_str = f"  [{', '.join(fmt)}]" if fmt else ""
                    lines.append(f"  {cell.coordinate}: {val_str!r}{fmt_str}")

        # Formulas — collect all, then summarize by column-runs
        formula_rows: list[tuple[str, int, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                if (cell.value is not None and isinstance(cell.value, str)
                        and cell.value.startswith("=")):
                    formula_rows.append((_col_letter(cell.column), cell.row, cell.value))
        if formula_rows:
            lines.append("")
            lines.append(f"Formulas ({len(formula_rows)} cells, grouped by column-run):")
            for summary in _summarize_column_runs(formula_rows):
                lines.append(f"  {summary}")

        # Charts
        if hasattr(ws, "_charts") and ws._charts:
            lines.append("")
            lines.append(f"Charts ({len(ws._charts)}):")
            for ch in ws._charts:
                ch_type = type(ch).__name__
                title = ""
                try:
                    if hasattr(ch, "title") and ch.title:
                        title = str(ch.title.tx.rich.p[0].r[0].t) if hasattr(ch.title, "tx") else str(ch.title)
                except Exception:
                    pass
                lines.append(f"  {ch_type}: {title or '(no title)'}")

        # Named tables
        if ws.tables:
            lines.append("")
            lines.append(f"Tables: {[(name, t.ref) for name, t in ws.tables.items()]}")

        lines.append("")

    return "\n".join(lines), cell_value_candidates


def build_scrub_terms(stem: str, audit: str, cell_values: list[str],
                      extra_names: list[str]) -> list[str]:
    """Names in xlsx submissions live in: file properties (already scrubbed by
    omission from the audit), cell values (Name: header cells), and the
    filename. We also scrub the operator-supplied .known_names.txt."""
    terms = set(extra_names)
    # Any cell value that looks name-like — alpha + spaces, 3-80 chars
    for v in cell_values:
        v = v.strip()
        if re.match(r"^[A-Za-z][A-Za-z\s.'\-]{2,80}$", v) and " " in v:
            terms.add(v)
    terms.update(EMAIL_RE.findall(audit))
    terms.update(USERPATH_RE.findall(audit))
    name_field = stem.split("_", 1)[0].lower()
    if len(name_field) >= 4 and name_field not in {"xlsx", "workbook", "submission"}:
        terms.add(name_field)
    return sorted((t for t in terms if t), key=len, reverse=True)


def scrub(text: str, terms: list[str]) -> tuple[str, int]:
    n = 0
    for t in terms:
        pat = re.compile(re.escape(t), re.IGNORECASE)
        text, k = pat.subn("[REDACTED]", text)
        n += k
    text, k1 = EMAIL_RE.subn("[REDACTED]", text)
    text, k2 = USERPATH_RE.subn("[REDACTED]", text)
    text, k3 = SECRET_PREFIX_RE.subn("[REDACTED-SECRET]", text)
    text, k4 = SECRET_ASSIGN_RE.subn(r"\1=[REDACTED-SECRET]", text)
    return text, n + k1 + k2 + k3 + k4


def main() -> int:
    ap = argparse.ArgumentParser(
        description="FERPA de-identify Excel (.xlsx) submissions via the "
                    "workbook-audit pattern. Generates a plain-text "
                    "description of structure/formulas/formatting/charts; "
                    "scrubs metadata + name-like cells; FERPA-safe.")
    ap.add_argument("--version", action="version", version=f"canvas-toolbox {__version__}")
    ap.add_argument("--challenge-dir", dest="challenge_dir", default=None,
                    help="Convention base path (e.g. grading/budget_lab).")
    ap.add_argument("--in", dest="indir", default=None,
                    help="Raw submissions directory (default: <challenge-dir>/submissions_raw)")
    ap.add_argument("--out", dest="outdir", default=None,
                    help="Output directory for keyed .md files (default: <challenge-dir>/submissions_deid)")
    ap.add_argument("--map", dest="mapfile", default=None,
                    help="Path to .keymap.json (default: <challenge-dir>/.keymap.json)")
    ap.add_argument("--names", dest="namesfile", default=None,
                    help="Optional gitignored file of known names (default: <challenge-dir>/.known_names.txt)")
    ap.add_argument("--prefix", default=None,
                    help="Key prefix (e.g. BUDGET). Default: uppercased basename of --challenge-dir.")
    args = ap.parse_args()

    if args.challenge_dir:
        cd = Path(args.challenge_dir)
        args.indir = args.indir or str(cd / "submissions_raw")
        args.outdir = args.outdir or str(cd / "submissions_deid")
        args.mapfile = args.mapfile or str(cd / ".keymap.json")
        args.namesfile = args.namesfile or str(cd / ".known_names.txt")
        args.prefix = args.prefix or cd.name.upper().replace("_", "-")

    missing = [a for a in ("indir", "outdir", "mapfile", "prefix") if not getattr(args, a)]
    if missing:
        print(f"Missing required arguments: {missing}. Pass --challenge-dir OR all of "
              f"--in/--out/--map/--prefix.", file=sys.stderr)
        return 1

    indir = Path(args.indir)
    outdir = Path(args.outdir)
    mapfile = Path(args.mapfile)
    outdir.mkdir(parents=True, exist_ok=True)

    extra_names: list[str] = []
    nf = Path(args.namesfile) if args.namesfile else None
    if nf and nf.exists():
        extra_names = [ln.strip() for ln in nf.read_text(encoding="utf-8").splitlines()
                       if ln.strip() and not ln.lstrip().startswith("#")]

    files = sorted(indir.glob("*.xlsx"))
    if not files:
        print(f"No .xlsx files in {indir}/ — nothing to do.", file=sys.stderr)
        return 1

    keymap: dict[str, str] = {}
    if mapfile.exists():
        try:
            keymap = json.loads(mapfile.read_text(encoding="utf-8")).get("map", {})
        except Exception:
            pass

    ok = fail = 0
    for f in files:
        key = key_for(f.name, args.prefix)
        try:
            audit, cell_values = build_audit(f)
            terms = build_scrub_terms(f.stem, audit, cell_values, extra_names)
            scrubbed, redactions = scrub(audit, terms)
            out = outdir / f"{key}.md"
            header = f"# Submission {key}\n\n_Source: Excel workbook (audit pattern; metadata not included)_\n\n"
            out.write_text(header + scrubbed + "\n", encoding="utf-8")
            keymap[key] = f.name
            ok += 1
            # FERPA: print KEY + counts only
            print(f"  {key}: {len(audit)} audit chars, {redactions} redactions -> {out.name}")
        except Exception as e:
            print(f"  {key}: SKIP — error ({type(e).__name__})")
            fail += 1

    mapfile.write_text(json.dumps(
        {"_warning": "FERPA — do NOT commit, do NOT let an AI read this. "
                     "Local re-identification only.",
         "map": keymap}, indent=2), encoding="utf-8")
    print(f"\n{ok} de-identified, {fail} skipped. "
          f"Map ({len(keymap)} keys) -> {mapfile} (gitignored, never read by AI).")
    return 0 if fail == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
